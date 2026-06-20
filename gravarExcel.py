import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)

# ── Paleta de cores ──────────────────────────────────────────────────────────
COR_HEADER_FUNDO   = "1F4E79"
COR_HEADER_FONTE   = "FFFFFF"
COR_LINHA_PAR      = "D6E4F0"
COR_LINHA_IMPAR    = "FFFFFF"
COR_AVISO_FUNDO    = "FFF2CC"
COR_AVISO_BORDA    = "F4B942"
COR_AVISO_FONTE    = "7B3F00"
COR_ALERTA_NUMERO  = "C00000"

# Cores para aviso de email ausente
COR_EMAIL_FUNDO    = "FCE4EC"   # rosa claro
COR_EMAIL_BORDA    = "C62828"   # vermelho escuro
COR_EMAIL_FONTE    = "B71C1C"   # vermelho profundo
COR_EMAIL_TITULO   = "FF1744"   # vermelho vivo


# ── Helpers de estilo ────────────────────────────────────────────────────────
def _borda(cor="BFBFBF", estilo="thin"):
    lado = Side(style=estilo, color=cor)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _borda_aviso():
    lado = Side(style="medium", color=COR_AVISO_BORDA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _borda_email():
    lado = Side(style="medium", color=COR_EMAIL_BORDA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _fill(hex_cor):
    return PatternFill("solid", fgColor=hex_cor)

def _font(bold=False, cor="000000", tamanho=11, nome="Arial"):
    return Font(name=nome, bold=bold, color=cor, size=tamanho)

def _alinhar(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Formatar cabeçalho ───────────────────────────────────────────────────────
def _formatar_cabecalho(ws):
    for cell in ws[1]:
        cell.font      = _font(bold=True, cor=COR_HEADER_FONTE, tamanho=11)
        cell.fill      = _fill(COR_HEADER_FUNDO)
        cell.alignment = _alinhar()
        cell.border    = _borda(cor="FFFFFF", estilo="thin")
    ws.row_dimensions[1].height = 28


# ── Formatar linha de dados ──────────────────────────────────────────────────
def _formatar_linha_dado(ws, row_num):
    cor = COR_LINHA_PAR if row_num % 2 == 0 else COR_LINHA_IMPAR
    for cell in ws[row_num]:
        cell.fill      = _fill(cor)
        cell.font      = _font(tamanho=10)
        cell.alignment = _alinhar()
        cell.border    = _borda()
    ws.row_dimensions[row_num].height = 22


# ── Aviso de conflito (amarelo) ──────────────────────────────────────────────
def _criar_textbox_aviso(ws, nome_disciplina, data, qtd_conflitos):
    row = ws.max_row + 2

    titulo_cell = ws.cell(row=row, column=1)
    titulo_cell.value     = "⚠️  AVISO DE CONFLITO"
    titulo_cell.font      = _font(bold=True, cor=COR_AVISO_BORDA, tamanho=11)
    titulo_cell.fill      = _fill("FFC000")
    titulo_cell.alignment = _alinhar()
    titulo_cell.border    = _borda_aviso()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 24

    row += 1
    corpo_cell = ws.cell(row=row, column=1)
    corpo_cell.value = (
        f"Curso: {nome_disciplina}\n"
        f"{data}: ⚠️ conflitos com {qtd_conflitos} alunos"
    )
    corpo_cell.font      = _font(
        bold=True,
        cor=COR_ALERTA_NUMERO if qtd_conflitos >= 10 else COR_AVISO_FONTE,
        tamanho=11
    )
    corpo_cell.fill      = _fill(COR_AVISO_FUNDO)
    corpo_cell.alignment = _alinhar(wrap=True)
    corpo_cell.border    = _borda_aviso()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 48


# ── Aviso de email ausente (vermelho) ────────────────────────────────────────
def _criar_textbox_sem_email(ws, professor, codigo, nome_disciplina, data_aula, modulo):
    """
    Bloco visual vermelho indicando que o professor não tem email cadastrado.
    Ocupa colunas A–F com fundo rosa e borda vermelha escura.
    """
    row = ws.max_row + 2

    # Linha de título
    titulo_cell = ws.cell(row=row, column=1)
    titulo_cell.value     = "🚫  SEM EMAIL INSTITUCIONAL"
    titulo_cell.font      = _font(bold=True, cor="FFFFFF", tamanho=11)
    titulo_cell.fill      = _fill(COR_EMAIL_BORDA)
    titulo_cell.alignment = _alinhar()
    titulo_cell.border    = _borda_email()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 24

    # Linha de corpo
    row += 1
    corpo_cell = ws.cell(row=row, column=1)
    corpo_cell.value = (
        f"Professor: {professor}\n"
        f"Disciplina: {codigo} – {nome_disciplina} | Módulo: {modulo} | Aula: {data_aula}\n"
        f"❌ Email institucional não cadastrado — notificação não enviada."
    )
    corpo_cell.font      = _font(bold=True, cor=COR_EMAIL_FONTE, tamanho=10)
    corpo_cell.fill      = _fill(COR_EMAIL_FUNDO)
    corpo_cell.alignment = _alinhar(h="left", wrap=True)
    corpo_cell.border    = _borda_email()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 56


# ── Ajustar largura das colunas automaticamente ──────────────────────────────
def _auto_ajustar_colunas(ws):
    larguras_minimas = {1: 20, 2: 30, 3: 14, 4: 16, 5: 30, 6: 14}
    for col in ws.columns:
        col_idx    = col[0].column
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value and not isinstance(cell, type(None))),
            default=10
        )
        ws.column_dimensions[col_letter].width = max(
            max_len + 4,
            larguras_minimas.get(col_idx, 12)
        )


def _congelar_cabecalho(ws):
    ws.freeze_panes = "A2"


def _garantir_cabecalho(wb, caminho_excel):
    """Cria workbook e cabeçalho se o arquivo ainda não existir."""
    ws = wb.active
    if not os.path.exists(caminho_excel):
        ws.title = "Conflitos"
        cabecalho = [
            "Nº Disciplina",
            "Nome Disciplina",
            "Data",
            "Módulo",
            "Disciplina Conflito",
            "Qtd Conflitos",
        ]
        ws.append(cabecalho)
        _formatar_cabecalho(ws)
        _congelar_cabecalho(ws)
    return ws


# ── Gravar conflito de horário ───────────────────────────────────────────────
def gravar_conflitos_excel(
    caminho_excel,
    numero_disciplina,
    nome_disciplina,
    data,
    modulo,
    conflitos
):
    novo_arquivo = not os.path.exists(caminho_excel)
    wb = Workbook() if novo_arquivo else load_workbook(caminho_excel)
    ws = _garantir_cabecalho(wb, caminho_excel) if novo_arquivo else wb.active

    disciplina_conflito = conflitos[0]["disciplina_conflito"] if conflitos else ""
    qtd_conflitos       = len(conflitos)

    ws.append([
        numero_disciplina,
        nome_disciplina,
        data,
        modulo,
        disciplina_conflito,
        qtd_conflitos,
    ])
    _formatar_linha_dado(ws, ws.max_row)

    celula_qtd = ws.cell(row=ws.max_row, column=6)
    if qtd_conflitos >= 10:
        celula_qtd.font = _font(bold=True, cor=COR_ALERTA_NUMERO, tamanho=11)

    _criar_textbox_aviso(ws, nome_disciplina, data, qtd_conflitos)
    _auto_ajustar_colunas(ws)
    wb.save(caminho_excel)

    print(
        f"📗   Salvo: {caminho_excel}\n"
        f"     Curso: {nome_disciplina} | {data} | {qtd_conflitos} conflito(s)"
    )


# ── Gravar aviso de professor sem email ─────────────────────────────────────
def gravar_sem_email_excel( caminho_excel, professor, codigo, nome_disciplina, data_aula, modulo):
    """
    Registra uma linha de dados + bloco visual vermelho no Excel
    indicando que o professor não possui email institucional cadastrado.
    """
    novo_arquivo = not os.path.exists(caminho_excel)
    wb = Workbook() if novo_arquivo else load_workbook(caminho_excel)
    ws = _garantir_cabecalho(wb, caminho_excel) if novo_arquivo else wb.active

    # Linha de dados com traço na coluna de conflitos (não se aplica)
    ws.append([
        codigo,
        nome_disciplina,
        professor,          # coluna "Data" reaproveitada para professor
        modulo,
        "—",               # sem disciplina conflito
        "—",               # sem qtd conflitos
    ])
    _formatar_linha_dado(ws, ws.max_row)

    # Destacar a linha inteira em rosa claro
    for cell in ws[ws.max_row]:
        cell.fill = _fill(COR_EMAIL_FUNDO)
        cell.font = _font(bold=True, cor=COR_EMAIL_FONTE, tamanho=10)

    _criar_textbox_sem_email(ws, professor, codigo, nome_disciplina, data_aula, modulo)
    _auto_ajustar_colunas(ws)
    wb.save(caminho_excel)

    print(
        f"📗    Salvo conflitos_horario.xlsx"
    )